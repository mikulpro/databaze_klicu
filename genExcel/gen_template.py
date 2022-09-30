from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

wb = Workbook()

border_thin = Side(border_style='thin', color='FF000000')
border_thin_full = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

alignment = Alignment(horizontal='center',
                        vertical='center',
                        text_rotation=0,
                        wrap_text=False,
                        shrink_to_fit=False,
                        indent=0)


dest_filename = 'template.xlsx'

ws = wb.active

ws.row_dimensions[1].height = 50
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 15

ws['A1'] = "Výpis evidence vypůjček klíčů"
ws['A1'].font = Font(name='Calibri', size=18)
ws['A1'].alignment = alignment
ws.merge_cells('A1:G1')

ws['A2'] = "Vypůjčení"
ws['A2'].border = border_thin_full
ws['A2'].alignment = alignment
ws.merge_cells('A2:B2')

ws['A3'] = "Datum"
ws['A3'].border = border_thin_full
ws['A3'].alignment = alignment

ws['B3'] = "Čas"
ws['B3'].border = border_thin_full
ws['B3'].alignment = alignment

ws['C2'] = "Číslo klíče"
ws['C2'].border = border_thin_full
ws['C2'].alignment = alignment
ws.merge_cells('C2:C3')

ws['D2'] = "Zapůjčení komu"
ws['D2'].border = border_thin_full
ws['D2'].alignment = alignment
ws.merge_cells('D2:D3')

ws['E2'] = "Vrácení"
ws['E2'].border = border_thin_full
ws['E2'].alignment = alignment
ws.merge_cells('E2:F2')

ws['E3'] = "Datum"
ws['E3'].border = border_thin_full
ws['E3'].alignment = alignment

ws['F3'] = "Čas"
ws['F3'].border = border_thin_full
ws['F3'].alignment = alignment

ws['G2'] = "Půjčovatel"
ws['G2'].border = border_thin_full
ws['G2'].alignment = alignment
ws.merge_cells('G2:G3')

wb.save(filename=dest_filename)