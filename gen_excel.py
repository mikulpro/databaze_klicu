from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment


class ExcelGenerator:
    border_thin = Side(border_style='thin', color='FF000000')
    border_thin_full = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_thin_sides = Border(left=border_thin, right=border_thin)
    border_thin_last = Border(left=border_thin, right=border_thin, bottom=border_thin)

    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)

    @staticmethod
    def gen_template(ws):
        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 18

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 15

        ws['A1'] = "Výpis evidence vypůjček klíčů"
        ws['A1'].font = Font(name='Calibri', size=18)
        ws['A1'].alignment = ExcelGenerator.alignment
        ws.merge_cells('A1:G1')

        ws['A2'] = "Vypůjčení"
        ws['A2'].border = ExcelGenerator.border_thin_full
        ws['A2'].alignment = ExcelGenerator.alignment
        ws.merge_cells('A2:B2')

        ws['A3'] = "Datum"
        ws['A3'].border = ExcelGenerator.border_thin_full
        ws['A3'].alignment = ExcelGenerator.alignment

        ws['B3'] = "Čas"
        ws['B3'].border = ExcelGenerator.border_thin_full
        ws['B3'].alignment = ExcelGenerator.alignment

        ws['C2'] = "Číslo klíče"
        ws['C2'].border = ExcelGenerator.border_thin_full
        ws['C2'].alignment = ExcelGenerator.alignment
        ws.merge_cells('C2:C3')

        ws['D2'] = "Zapůjčení komu"
        ws['D2'].border = ExcelGenerator.border_thin_full
        ws['D2'].alignment = ExcelGenerator.alignment
        ws.merge_cells('D2:D3')

        ws['E2'] = "Vrácení"
        ws['E2'].border = ExcelGenerator.border_thin_full
        ws['E2'].alignment = ExcelGenerator.alignment
        ws.merge_cells('E2:F2')

        ws['E3'] = "Datum"
        ws['E3'].border = ExcelGenerator.border_thin_full
        ws['E3'].alignment = ExcelGenerator.alignment

        ws['F3'] = "Čas"
        ws['F3'].border = ExcelGenerator.border_thin_full
        ws['F3'].alignment = ExcelGenerator.alignment

        ws['G2'] = "Půjčovatel"
        ws['G2'].border = ExcelGenerator.border_thin_full
        ws['G2'].alignment = ExcelGenerator.alignment
        ws.merge_cells('G2:G3')

    @staticmethod
    def append_data(ws, data):
        for row in range(len(data)):
            if row == len(data)-1:
                border = ExcelGenerator.border_thin_last
            else:
                border = ExcelGenerator.border_thin_sides
            for col in range(len(data[row])):
                cell_coord = chr(65+col) + str(4+row)
                ws[cell_coord] = data[row][col]
                ws[cell_coord].border = border

    @staticmethod
    def gen_excel_file(data, dest_filename="export.xlsx"):
        wb = Workbook()
        ws = wb.active
        ExcelGenerator.gen_template(ws)
        ExcelGenerator.append_data(ws, data)
        wb.save(filename=dest_filename)


if __name__ == '__main__':
    data = [
        ("10.10.2022", "14:53", "CP1.03", "Jiří Fišer", "10.10.2022", "16:48", "Novák"),
        ("10.10.2022", "15:56", "CP1.04", "Jan Krejčí", "10.10.2022", "18:46", "Novák"),
        ("10.10.2022", "17:53", "CP1.03", "Petr Haberzettl", "10.10.2022", "19:48", "Novák"),
    ]
    ExcelGenerator.gen_excel_file(data)
    print('Excel file was sucessfully generated "export.xlsx"')
